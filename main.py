import re
import pathlib
import logging
import datetime as dt
import time
import requests
from typing import List, Dict, Optional, Tuple
import traceback
import tkinter as tk
from tkinter import ttk, scrolledtext
import threading
import fitz
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',
                   handlers=[logging.FileHandler('logs/nfe_processor.log', encoding='utf-8')])
logger = logging.getLogger(__name__)

class TextHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget

    def emit(self, record):
        msg = self.format(record)
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, msg + '\n')
        self.text_widget.see(tk.END)
        self.text_widget.configure(state='disabled')

class NFeProcessorSimplified:
    def __init__(self, log_widget=None):
        self.base_dir = pathlib.Path(__file__).parent
        self.input_dir = self.base_dir / "input"
        self.output_dir = self.base_dir / "output" 
        self.processed_dir = self.base_dir / "processed"
        self.logs_dir = self.base_dir / "logs"
        self.sheets_file = self.base_dir / "Sheets.xlsx"
        
        for dir_path in [self.input_dir, self.output_dir, self.processed_dir, self.logs_dir]:
            dir_path.mkdir(exist_ok=True)
        
        self.cache_cnpj: Dict[str, Optional[Dict]] = {}
        self._setup_patterns()
        
        if log_widget:
            text_handler = TextHandler(log_widget)
            text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
            logger.addHandler(text_handler)
    
    def _setup_patterns(self):
        self.patterns = {
            'decimal_trans': str.maketrans({".": "", ",": "."}),
            'cnpj': re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}"),
            'data_emissao': re.compile(r"EMISS[ÃA]O[:\s]*([0-9]{2}/[0-9]{2}/[0-9]{4})", re.I),
            'numero_nfe': re.compile(r"NF-e\s+N[ºº°]\s*(\d{1,9})", re.I),
            'numero_simples': re.compile(r"N[ºº°]\s*(\d+)", re.I),
        }
    
    def _validar_cnpj(self, cnpj: str) -> bool:
        if len(cnpj) != 14 or not cnpj.isdigit():
            return False
            
        def calc_digito(cnpj: str, pesos: List[int]) -> int:
            soma = sum(int(cnpj[i]) * pesos[i] for i in range(len(pesos)))
            resto = soma % 11
            return 0 if resto < 2 else 11 - resto
            
        pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        pesos2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        
        return (int(cnpj[12]) == calc_digito(cnpj, pesos1) and 
                int(cnpj[13]) == calc_digito(cnpj, pesos2))
    
    def consultar_cnpj_api(self, cnpj: str) -> Optional[Dict]:
        if not cnpj:
            return None
            
        cnpj_limpo = re.sub(r'[^\d]', '', cnpj)
        
        if not self._validar_cnpj(cnpj_limpo):
            return None
            
        if cnpj_limpo in self.cache_cnpj:
            return self.cache_cnpj[cnpj_limpo]
            
        try:
            time.sleep(0.3)
            response = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_limpo}", timeout=20)
            
            if response.status_code == 200:
                dados = response.json()
                resultado = {'razao_social': dados.get('razao_social', '').strip().upper()}
                self.cache_cnpj[cnpj_limpo] = resultado
                return resultado
            else:
                self.cache_cnpj[cnpj_limpo] = None
                
        except Exception as e:
            logger.error(f"Erro ao consultar CNPJ {cnpj_limpo}: {e}")
            self.cache_cnpj[cnpj_limpo] = None
            
        return None
    
    def identificar_tipo_material(self, descricao: str) -> Optional[str]:
        if not descricao:
            return None
            
        desc_lower = descricao.lower()
        
        if any(palavra in desc_lower for palavra in [
            'plastico', 'plástico', 'pet', 'pvc', 'pead', 'pebd', 'pp', 'ps',
            'polietileno', 'polipropileno', 'poliestireno'
        ]):
            return 'PLASTICO'
        
        elif any(palavra in desc_lower for palavra in [
            'metal', 'ferro', 'aco', 'aço', 'ferroso', 'inox', 'inoxidavel', 
            'aluminio', 'alumínio', 'cobre', 'bronze', 'latao', 'latão', 
            'zinco', 'chumbo', 'sucata metalica', 'sucata metálica'
        ]):
            return 'METAL'
        
        elif any(palavra in desc_lower for palavra in ['vidro', 'cristal', 'garrafa vidro']):
            return 'VIDRO'
        
        elif any(palavra in desc_lower for palavra in ['papel', 'papelao', 'papelão', 'cartao', 'cartão']):
            return 'PAPEL'
        
        return None
    
    def to_float(self, text: str) -> Optional[float]:
        if not text or not isinstance(text, str):
            return None
        
        try:
            clean_text = re.sub(r'[^\d,.]', '', text.strip())
            if not clean_text:
                return None
            return float(clean_text.translate(self.patterns['decimal_trans']))
        except (ValueError, TypeError):
            return None
    
    def extract_pdf_text(self, pdf_path: pathlib.Path) -> str:
        try:
            with fitz.open(pdf_path) as doc:
                return "\n".join(page.get_text("text") for page in doc)
        except Exception as e:
            logger.error(f"Erro ao extrair texto do PDF {pdf_path}: {e}")
            return ""
    
    def extract_cnpjs(self, text: str) -> Tuple[str, str]:
        cnpjs = []
        patterns = [
            r'CNPJ\s*/\s*CPF[:\s]*(\d{2}\.?\d{3}\.?\d{3}\/?\d{4}-?\d{2})',
            r'(\d{2}\.?\d{3}\.?\d{3}\/?\d{4}-?\d{2})',
            r'(\d{14})'
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                cnpj_limpo = re.sub(r'[^\d]', '', match)
                if len(cnpj_limpo) == 14 and self._validar_cnpj(cnpj_limpo):
                    cnpj_fmt = f"{cnpj_limpo[:2]}.{cnpj_limpo[2:5]}.{cnpj_limpo[5:8]}/{cnpj_limpo[8:12]}-{cnpj_limpo[12:14]}"
                    if cnpj_fmt not in cnpjs:
                        cnpjs.append(cnpj_fmt)
        
        return (cnpjs[0] if cnpjs else "", cnpjs[1] if len(cnpjs) > 1 else "")
    
    def extract_metadata(self, text: str) -> Dict[str, any]:
        metadata = {
            "numero_nfe": "",
            "data_emissao": None,
            "emit_razao_social": "",
            "emit_cnpj": "",
            "dest_razao_social": "",
            "dest_cnpj": ""
        }
        
        nfe_match = self.patterns['numero_nfe'].search(text)
        if nfe_match:
            metadata["numero_nfe"] = nfe_match.group(1)
        else:
            num_match = self.patterns['numero_simples'].search(text)
            if num_match:
                metadata["numero_nfe"] = num_match.group(1)
        
        data_match = self.patterns['data_emissao'].search(text)
        if data_match:
            try:
                metadata["data_emissao"] = dt.datetime.strptime(data_match.group(1), "%d/%m/%Y").date()
            except ValueError:
                pass
        
        cnpj_emit, cnpj_dest = self.extract_cnpjs(text)
        metadata["emit_cnpj"] = cnpj_emit
        metadata["dest_cnpj"] = cnpj_dest
        
        if cnpj_emit:
            dados_emit = self.consultar_cnpj_api(cnpj_emit)
            if dados_emit:
                metadata["emit_razao_social"] = dados_emit.get('razao_social', '')
        
        if cnpj_dest:
            dados_dest = self.consultar_cnpj_api(cnpj_dest)
            if dados_dest:
                metadata["dest_razao_social"] = dados_dest.get('razao_social', '')
        
        return metadata
    
    def extract_items_pdfplumber(self, pdf_path: pathlib.Path) -> List[Dict]:
        items = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    
                    for table in tables:
                        if not table:
                            continue
                        
                        header_found = False
                        for row in table:
                            if not row:
                                continue
                            
                            if not header_found and row[0] and "PROD" in str(row[0]).upper():
                                header_found = True
                                continue
                            
                            if not header_found:
                                continue
                            
                            if (len(row) >= 9 and row[2] and 
                                re.fullmatch(r"\d{8}", str(row[2]).strip())):
                                
                                descricao = str(row[1] or "").strip()
                                quantidade = self.to_float(str(row[6] or "").strip())
                                valor_total = self.to_float(str(row[8] or "").strip())
                                
                                tipo_material = self.identificar_tipo_material(descricao)
                                
                                if tipo_material:
                                    item = {
                                        'descricao': descricao,
                                        'quantidade': quantidade,
                                        'valor': valor_total,
                                        'tipo_material': tipo_material
                                    }
                                    items.append(item)
        
        except Exception as e:
            logger.error(f"Erro ao extrair itens com pdfplumber: {e}")
        
        return items
    
    def extract_items_regex(self, text: str) -> List[Dict]:
        items = []
        
        item_pattern = re.compile(
            r"(?P<codigo_item>\d{3})\s+"
            r"(?P<descricao>.+?)\s+"
            r"(?P<ncm>\d{8})\s+"
            r"(?P<cst>\d{3})\s+"
            r"(?P<cfop>\d{4})\s+"
            r"(?P<unid>[A-Z]{2,4})\s+"
            r"(?P<quantidade>[0-9\.\,]+)\s+"
            r"(?P<valor_unit>[0-9\.\,]+)\s+"
            r"(?P<valor_total>[0-9\.\,]+)", re.S
        )
        
        for match in item_pattern.finditer(text):
            data = match.groupdict()
            tipo_material = self.identificar_tipo_material(data.get('descricao', ''))
            
            if tipo_material:
                item = {
                    'descricao': data['descricao'],
                    'quantidade': self.to_float(data['quantidade']),
                    'valor': self.to_float(data['valor_total']),
                    'tipo_material': tipo_material
                }
                items.append(item)
        
        return items
    
    def process_pdf(self, pdf_path: pathlib.Path) -> List[Dict]:
        logger.info(f"Processando: {pdf_path.name}")
        
        try:
            text = self.extract_pdf_text(pdf_path)
            if not text:
                return []
            
            metadata = self.extract_metadata(text)
            items = self.extract_items_pdfplumber(pdf_path)
            if not items:
                items = self.extract_items_regex(text)
            
            for item in items:
                item.update(metadata)
            
            if items:
                tipos = set(item.get('tipo_material', '') for item in items)
                logger.info(f"✓ {pdf_path.name} - {len(items)} itens ({', '.join(tipos)})")
            
            return items
            
        except Exception as e:
            logger.error(f"❌ Erro ao processar {pdf_path.name}: {e}")
            return []
    
    def process_all_pdfs(self) -> Tuple[List[Dict], List[str]]:
        all_items = []
        failed_files = []
        
        pdf_files = sorted(self.input_dir.glob("*.pdf"))
        
        if not pdf_files:
            logger.warning("Nenhum arquivo PDF encontrado na pasta 'input'")
            return all_items, failed_files
        
        logger.info(f"Processando {len(pdf_files)} arquivos PDF")
        
        for pdf_path in pdf_files:
            try:
                items = self.process_pdf(pdf_path)
                if items:
                    all_items.extend(items)
                    processed_path = self.processed_dir / pdf_path.name
                    pdf_path.rename(processed_path)
                else:
                    failed_files.append(pdf_path.name)
            except Exception as e:
                failed_files.append(pdf_path.name)
                logger.error(f"Falha ao processar {pdf_path.name}: {e}")
        
        return all_items, failed_files
    
    def save_to_sheets(self, items: List[Dict]) -> pathlib.Path:
        try:
            df = pd.DataFrame(items)
            
            columns_order = [
                "emit_razao_social", "emit_cnpj", "dest_razao_social", "dest_cnpj",
                "numero_nfe", "data_emissao", "quantidade", "valor", "tipo_material", "descricao"
            ]
            
            existing_cols = [col for col in columns_order if col in df.columns]
            df = df[existing_cols]
            
            df = df.rename(columns={
                'emit_razao_social': 'Razão Social Emitente',
                'emit_cnpj': 'CNPJ Emitente',
                'dest_razao_social': 'Razão Social Destinatário', 
                'dest_cnpj': 'CNPJ Destinatário',
                'numero_nfe': 'Número NFe',
                'data_emissao': 'Data Emissão',
                'quantidade': 'Quantidade',
                'valor': 'Valor Total',
                'tipo_material': 'Tipo Material',
                'descricao': 'Descrição'
            })
            
            if self.sheets_file.exists():
                try:
                    with pd.ExcelWriter(self.sheets_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='lancamentos_nf', index=False)
                        
                        resumo = df.groupby('Tipo Material').agg({
                            'Quantidade': 'sum',
                            'Valor Total': 'sum',
                            'Número NFe': 'count'
                        }).rename(columns={'Número NFe': 'Qtd Registros'})
                        resumo.to_excel(writer, sheet_name='Resumo_por_Material')
                    
                    wb = load_workbook(self.sheets_file)
                    if 'lancamentos_nf' in wb.sheetnames:
                        worksheet = wb['lancamentos_nf']
                        for column in worksheet.columns:
                            max_length = max(len(str(cell.value)) for cell in column if cell.value)
                            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
                        wb.save(self.sheets_file)
                    
                except Exception:
                    with pd.ExcelWriter(self.sheets_file, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='lancamentos_nf', index=False)
                        
                        resumo = df.groupby('Tipo Material').agg({
                            'Quantidade': 'sum',
                            'Valor Total': 'sum',
                            'Número NFe': 'count'
                        }).rename(columns={'Número NFe': 'Qtd Registros'})
                        resumo.to_excel(writer, sheet_name='Resumo_por_Material')
            else:
                with pd.ExcelWriter(self.sheets_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='lancamentos_nf', index=False)
                    
                    resumo = df.groupby('Tipo Material').agg({
                        'Quantidade': 'sum',
                        'Valor Total': 'sum',
                        'Número NFe': 'count'
                    }).rename(columns={'Número NFe': 'Qtd Registros'})
                    resumo.to_excel(writer, sheet_name='Resumo_por_Material')
                
                wb = load_workbook(self.sheets_file)
                if 'lancamentos_nf' in wb.sheetnames:
                    worksheet = wb['lancamentos_nf']
                    for column in worksheet.columns:
                        max_length = max(len(str(cell.value)) for cell in column if cell.value)
                        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
                    wb.save(self.sheets_file)
            
            logger.info(f"✔ Dados salvos em: {self.sheets_file}")
            return self.sheets_file
            
        except Exception as e:
            logger.error(f"Erro ao salvar dados: {e}")
            raise
    
    def run(self):
        logger.info("INICIANDO PROCESSAMENTO - MATERIAIS RECICLÁVEIS")
        
        all_items, failed_files = self.process_all_pdfs()
        
        if all_items:
            output_file = self.save_to_sheets(all_items)
            logger.info(f"✅ {len(all_items)} itens processados - {output_file}")
            
            df = pd.DataFrame(all_items)
            materiais = df['tipo_material'].value_counts()
            valor_total = df['valor'].sum() if 'valor' in df.columns else 0
            
            logger.info(f"Valor total: R$ {valor_total:,.2f}")
            for material, qtd in materiais.items():
                logger.info(f"{material}: {qtd} itens")
        else:
            logger.warning("❌ Nenhum material encontrado")
        
        if failed_files:
            logger.warning(f"Arquivos sem materiais: {', '.join(failed_files)}")
        
        logger.info("Processamento concluído!")

class NFeProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("NFe Processor - Materiais Recicláveis")
        self.root.geometry("600x400")
        
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        ttk.Label(self.main_frame, text="Processador de Notas Fiscais (NF-e)", 
                 font=("Helvetica", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(self.main_frame, height=15, width=70, state='disabled')
        self.log_text.grid(row=1, column=0, columnspan=2, pady=10)
        
        self.process_button = ttk.Button(self.main_frame, text="Processar PDFs", command=self.start_processing)
        self.process_button.grid(row=2, column=0, pady=5, sticky=tk.W)
        
        ttk.Button(self.main_frame, text="Sair", command=self.root.quit).grid(row=2, column=1, pady=5, sticky=tk.E)
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.rowconfigure(1, weight=1)
        
        self.is_processing = False
    
    def start_processing(self):
        if self.is_processing:
            return
        
        self.is_processing = True
        self.process_button.configure(state='disabled')
        
        self.log_text.configure(state='normal')
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state='disabled')
        
        threading.Thread(target=self.run_processing, daemon=True).start()
    
    def run_processing(self):
        try:
            processor = NFeProcessorSimplified(self.log_text)
            processor.run()
        except Exception as e:
            logger.error(f"Erro fatal: {e}")
        finally:
            self.is_processing = False
            self.process_button.configure(state='normal')

def main():
    root = tk.Tk()
    app = NFeProcessorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()